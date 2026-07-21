import pandas as pd

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

from config import MAPPING
from project_rules import PROJECT_RULES


# =====================================================
# Ambil value berdasarkan alamat cell
# =====================================================
def get_cell(ws, cell):

    return ws[cell].value


# =====================================================
# Cari value berdasarkan label
#
# occurrence
# 1 = label pertama
# 2 = label kedua
# =====================================================
def get_label_value(ws, label, occurrence=1):

    label = label.strip().lower()

    found_count = 0

    for row in ws.iter_rows():

        for cell in row:

            if isinstance(cell, MergedCell):
                continue

            if cell.value is None:
                continue

            text = str(cell.value).strip().lower()

            if label not in text:
                continue

            found_count += 1

            if found_count != occurrence:
                continue

            start_row = cell.row
            start_col = cell.column

            # ============================================
            # PRIORITAS 1
            # Cari angka di kanan label
            # ============================================

            for c in range(start_col + 1, start_col + 11):

                value = ws.cell(start_row, c).value

                if isinstance(value, (int, float)):
                    return value

            # ============================================
            # PRIORITAS 2
            # Cari angka di bawah label
            # ============================================

            for r in range(start_row + 1, start_row + 6):

                value = ws.cell(r, start_col).value

                if isinstance(value, (int, float)):
                    return value

            # ============================================
            # PRIORITAS 3
            # Scan area sekitar
            # ============================================

            for r in range(start_row, start_row + 5):

                for c in range(start_col, start_col + 10):

                    value = ws.cell(r, c).value

                    if isinstance(value, (int, float)):
                        return value

    return None


# =====================================================
# Main Process
# =====================================================
def get_without_regional_cost(ws):

    start_row = None
    start_col = None

    for row in ws.iter_rows():

        for cell in row:

            if cell.value is None:
                continue

            text = str(cell.value).strip().lower()

            if "without regional cost" in text:

                start_row = cell.row
                start_col = cell.column
                break

        if start_row is not None:
            break

    if start_row is None:
        return {}

    values = []

    for r in range(start_row + 1, start_row + 20):

        value = ws.cell(row=r, column=start_col).value

        if isinstance(value, (int, float)):
            values.append(value)

    if len(values) < 2:
        return {}

    return {
        "Total Revenue": values[0],
        "Total Cost (Before Finance Cost)": values[1],
    }

def process_files(uploaded_files, progress_bar=None, status_text=None):

    hasil_semua = []
    error_log = []

    total_file = len(uploaded_files)

    for i, file in enumerate(uploaded_files, start=1):

        try:

            wb = load_workbook(
                file,
                data_only=True
            )

            sheet = wb.sheetnames[0]
            ws = wb[sheet]

            project_code = get_cell(ws, "J2")

            summary_type = (
                PROJECT_RULES
                .get(project_code, {})
                .get("summary_type")
            )

            special_summary = {}

            if summary_type == "without_regional_cost":
                special_summary = get_without_regional_cost(ws)


            hasil = {
                "File Name": file.name,
                "Sheet Name": sheet,
                "PROJECT CODE": project_code
            }

            for field, config in MAPPING.items():

                if field in special_summary:
                    hasil[field] = special_summary[field]
                    continue

                try:

                    if config["type"] == "cell":

                        hasil[field] = get_cell(
                            ws,
                            config["value"]
                        )

                    elif config["type"] == "label":

                        labels = config["value"]

                        if isinstance(labels, str):
                            labels = [labels]

                        # ====================================
                        # Default ambil label pertama
                        # ====================================

                        occurrence = (
                            PROJECT_RULES
                            .get(project_code, {})
                            .get("occurrence", {})
                            .get(field, 1)
                        )

                        value = None

                        for lbl in labels:

                            value = get_label_value(
                                ws,
                                lbl,
                                occurrence
                            )

                            if value is not None:
                                break

                        hasil[field] = value

                    else:

                        hasil[field] = None

                except Exception:

                    hasil[field] = None

            hasil_semua.append(hasil)

        except Exception as e:

            error_log.append({
                "File Name": file.name,
                "Error": str(e)
            })

        # ============================================
        # Progress
        # ============================================

        if progress_bar is not None:
            progress_bar.progress(i / total_file)

        if status_text is not None:
            status_text.info(
                f"📄 Processing {i}/{total_file} : {file.name}"
            )

    hasil_df = pd.DataFrame(hasil_semua)
    error_df = pd.DataFrame(error_log)

    pd.set_option(
        "display.float_format",
        "{:,.2f}".format
    )

    return hasil_df, error_df
