import streamlit as st
import pandas as pd
from extractor import process_files
from io import BytesIO
from openpyxl import load_workbook

st.set_page_config(
    page_title="Excel Summary Extractor",
    page_icon="📊",
    layout="wide"
)

st.title("📊 Excel Summary Extractor")

st.write(
    "Upload satu atau beberapa file Excel kemudian klik **Process Files**."
)

uploaded_files = st.file_uploader(
    "Upload Excel Files",
    type=["xlsx"],
    accept_multiple_files=True
)

# ==========================
# FILE INFO
# ==========================

if uploaded_files:

    st.success(f"✅ {len(uploaded_files)} file berhasil dipilih")

    with st.expander("Detail File"):

        for f in uploaded_files:

            st.write(
                f"📄 {f.name} ({f.size/1024:.2f} KB)"
            )

    # ==========================
    # PROCESS BUTTON
    # ==========================

    if st.button("🚀 Process Files", use_container_width=True):

        progress_bar = st.progress(0)

        status_text = st.empty()

        with st.spinner("Sedang memproses file..."):

            hasil_df, error_df = process_files(
                uploaded_files,
                progress_bar=progress_bar,
                status_text=status_text
            )

        progress_bar.empty()
        status_text.empty()

        # ==========================
        # METRICS
        # ==========================

        total = len(uploaded_files)
        success = len(hasil_df)
        failed = len(error_df)

        col1, col2, col3 = st.columns(3)

        col1.metric("📂 Total File", total)
        col2.metric("✅ Success", success)
        col3.metric("❌ Error", failed)

        # ==========================
        # PREVIEW
        # ==========================

        st.subheader("Preview")

        preview_df = hasil_df.copy()

        numeric_cols = preview_df.select_dtypes(include="number").columns

        for col in numeric_cols:

            preview_df[col] = preview_df[col].map(
                lambda x: f"{x:,.2f}" if pd.notnull(x) else ""
            )

        st.dataframe(
            preview_df,
            use_container_width=True
        )

        # ==========================
        # DOWNLOAD
        # ==========================

        output = BytesIO()

        with pd.ExcelWriter(
            output,
            engine="openpyxl"
        ) as writer:

            hasil_df.to_excel(
                writer,
                index=False,
                sheet_name="Summary"
            )

        output.seek(0)

        wb = load_workbook(output)

        ws = wb["Summary"]

        for col in ws.iter_cols(min_row=2):

            for cell in col:

                if isinstance(cell.value, (int, float)):
                    cell.number_format = "#,##0.00"

        final_output = BytesIO()

        wb.save(final_output)

        final_output.seek(0)

        st.download_button(
            "⬇ Download Result",
            data=final_output,
            file_name="Summary_Output.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )

        # ==========================
        # ERROR LOG
        # ==========================

        st.subheader("⚠ Error Log")

        if error_df.empty:

            st.success("Tidak ada file yang error.")

        else:

            st.error(
                f"{len(error_df)} file gagal diproses."
            )

            st.dataframe(
                error_df,
                use_container_width=True
            )
